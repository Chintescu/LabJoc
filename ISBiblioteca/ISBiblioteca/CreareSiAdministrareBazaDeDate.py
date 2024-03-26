import openpyxl

class AdministrareBazaDate:

    def __init__(self, cale_baza_date):
        self.cale = cale_baza_date

    def Modificare_In_Baza_Date(self, Id, atribut, valuare):
        lin = self.Cautare_Dupa_Id(Id)
        col = self.CautareAtribut(atribut)
        if lin == -1 or col == -1:
            exit(0)
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        celula = foaie.cell(row=lin, column=col)
        celula.value = valuare
        bazaDate.save(self.cale)
        
    def Modificare_Valoare_In_Baza_Date(self, atribut, valuare):
        lin = self.CautareLinie(atribut)
        col = self.CautareAtribut(atribut)
        if  col == -1:
            exit(0)
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        print("linie ",lin)
        print("coloana ",col)
        celula = foaie.cell(row=lin, column=col)
        celula.value = valuare
        bazaDate.save(self.cale)
    def Scriere_In_Baza_Date(self, lin, col, valuare):
        if lin == -1 or col == -1:
            exit(0)
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        celula = foaie.cell(row=lin, column=col)
        celula.value = valuare
        bazaDate.save(self.cale)

    def Citire_Din_Baza_Date(self,lin,col):
        if lin == -1 or col == -1:
            exit(0)
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        celula = foaie.cell(row=lin, column=col)
        valuare = celula.value
        return valuare

    def Dimensiune_Baza_Date(self):
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        return [foaie.max_row,foaie.max_column]

    def Stergere_Inregistrare_Din_Baza_Date(self, Id):
        dimensiune = self.Dimensiune_Baza_Date()
        lin = self.Cautare_Dupa_Id(Id)
        if lin == -1:
            exit(0)
        for i in range(dimensiune[1]):
            self.Scriere_In_Baza_Date(lin, i + 1, None)

    def Stergere_Atribut_Inregistrare(self, Id, atribut):
        if self.CautareAtribut(atribut) == 1:
            print("Stergerea Id-ului va sterge intreaga inregistrare")
            self.Stergere_Inregistrare_Din_Baza_Date(Id)
        else:
            self.Modificare_In_Baza_Date(Id, atribut, None)
    
    def Cautare_Dupa_Id(self, Id):
        dimensiune = self.Dimensiune_Baza_Date()
        for i in range(dimensiune[0]):
            valuare = self.Citire_Din_Baza_Date(i+2, 1)
            if valuare == Id:
                return i + 2
        print("Id-ul: ", Id, " nu a fost gasit")
        return -1
        
    def CautareLinie(self, atribut):
        dimensiune = self.Dimensiune_Baza_Date()
        for i in range(dimensiune[0]):
            for j in range(dimensiune[0]):
                valuare = self.Citire_Din_Baza_Date(i+1, j+1)
                if valuare == atribut:
                    return i + 1
        print("atributul-ul: ", atribut, " nu a fost gasit")
        return -1
    
    def CautareAtribut(self, atribut):
        dimensiune = self.Dimensiune_Baza_Date()
        for i in range(dimensiune[0]):
            for j in range(dimensiune[0]):
                valuare = self.Citire_Din_Baza_Date(i+1, j+1)
                if valuare == atribut:
                    print("atributul-ul: ", atribut, " a fost gasit la coloana ",j+2," si linia ",i )
                    return j + 1
        print("atributul-ul: ", atribut, " nu a fost gasit")
        return -1
        
    def CautareAtributRow(self, atribut):
        dimensiune = self.Dimensiune_Baza_Date()
        for i in range(dimensiune[0]):
            for j in range(dimensiune[0]):
                valuare = self.Citire_Din_Baza_Date(i+1, j+1)
                if valuare == atribut:
                    print("atributul-ul: ", atribut, " a fost gasit la coloana ",j+2," si linia ",i )
                    return i+1
        print("atributul-ul: ", atribut, " nu a fost gasit")
        return -1
        
    def Adaugare_Inregistrare_Noua(self, lista_atribute):
        dimensiune = self.Dimensiune_Baza_Date()
        lin = 0
        for i in range(dimensiune[0]):
            if lista_atribute[0] == self.Citire_Din_Baza_Date(i+1, 1):
                print("Id-ul", lista_atribute[0], "nu este unic")
                exit(0)
            if self.Citire_Din_Baza_Date(i+1, 1) == None:
                lin = i+1
        if lin > 0:
            for i in range(dimensiune[1]):
                self.Scriere_In_Baza_Date(lin, i+1, lista_atribute[i])
        else:
            for i in range(dimensiune[1]):
                self.Scriere_In_Baza_Date(dimensiune[0]+1, i+1, lista_atribute[i])

    def CautareStoc(self, atribut):
        dimensiune = self.Dimensiune_Baza_Date()
        for i in range(dimensiune[0]):
            for j in range(dimensiune[0]):
                valuare = self.Citire_Din_Baza_Date(i+1, j+1)
                if valuare == atribut:
                    print("atributul-ul: ", atribut, " a fost gasit la coloana ",j+1," si linia ",i )
                    return self.Citire_Din_Baza_Date(i+1,j+2) 
        print("atributul-ul: ", atribut, " nu a fost gasit")
        return -1
    
    def EditareAtribut(self, valuare, line, collom):
        bazaDate = openpyxl.load_workbook(self.cale)
        foaie = bazaDate.active
        print("linie ",line)
        print("coloana ",collom)
        celula = foaie.cell(row=line, column=collom)
        celula.value = valuare
        bazaDate.save(self.cale)
            
        return 2
    
    
        