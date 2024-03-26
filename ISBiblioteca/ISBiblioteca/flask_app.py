import requests
from flask import (Flask, redirect, url_for,
                   render_template, request, session, jsonify)
from datetime import timedelta
from requests.auth import HTTPBasicAuth
from CreareSiAdministrareBazaDeDate import AdministrareBazaDate
import random
import AdministrareMail as mail
from openpyxl import load_workbook
import re
from flask_wtf import FlaskForm
from wtforms import StringField
from wtforms.validators import InputRequired
from GenericModel import GenericModel


class MyBooks:
    books_value = 0


class Email:
    email_receiver = ''


booksSize = MyBooks()
email_sender = 'zorbogdan2001@gmail.com'
email_password = 'attfdvpkuodyxkol'
email_receiver = ''
searchedAtribut = -2
cale = "database.xlsx"
caleCarti = "booksDatabase.xlsx"
book = AdministrareBazaDate(caleCarti)
elements = []
bd = AdministrareBazaDate(cale)
app = Flask(__name__)
app.secret_key = "hello"
app.permanent_session_lifetime = timedelta(minutes=5)

value = 0


class BasicForm(FlaskForm):
    ids = StringField("Book", validators=[InputRequired()])


@app.route("/")
def home():
    return render_template("index.html")


@app.route('/login', methods=['GET', 'POST'])
def login():
    msg = ''
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form:
        username = request.form['username']
        password = request.form['password']

        if bd.CautareAtribut(username) != -1 and bd.CautareAtribut(password) != -1:

            return redirect(url_for('home'))

        else:
            msg = 'Incorrect username / password !'
    return render_template('login.html', msg=msg)


@app.route('/register', methods=['GET', 'POST'])
def register():
    msg = ''
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form and 'email' in request.form:
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        number = random.randint(0, 5000)

        if bd.CautareAtribut(email) != -1:
            msg = 'Account already exists !'

        elif not re.match(r'[A-Za-z0-9]+', username):
            msg = 'Username must contain only characters and numbers !'
        elif not re.match(r'[^@]+@[^@]+\.[^@]+', email):
            msg = 'Invalid email address !'
        elif not username or not password or not email:
            msg = 'Please fill out the form !'
        else:
            Email.email_receiver = request.form['email']
            mail.Trimitere_Email(email_sender, email_password, 'Bun venit pe site', Email.email_receiver)
            inreg1 = [number, request.form['username'], request.form['password'], request.form['email']]
            bd.Adaugare_Inregistrare_Noua(inreg1)
            msg = 'You have successfully registered !'
    elif request.method == 'POST':
        msg = 'Please fill out the form !'
    return render_template('register.html', msg=msg)


@app.route('/delete', methods=['GET', 'POST'])
def delete():
    msg = ''
    if request.method == 'POST' and 'id' in request.form:
        id = request.form['id']
        k = 0
        line1 = ""
        for i in id:
            if i.isspace():
                print(i)
            else:
                line1 += i
        h = ord(line1)
        while h > 48:
            k += 1
            h -= 1
        book.Stergere_Inregistrare_Din_Baza_Date(k)
    return render_template('delete.html', msg=msg)


@app.route("/user")
def user():
    if "user" in session:
        user = session["user"]
        return f"<h1>{user}</h1>"
    else:
        return redirect(url_for("login"))


@app.route("/edit", methods=['GET', 'POST'])
def edit():
    bookEdit = load_workbook("booksDatabase.xlsx")
    sheet = bookEdit.active
    if request.method == 'POST':
        boom = request.form['book']
        nume = request.form['username']
        print(nume)
        name = ""
        line = ""
        collomn = ""
        j = 0
        for i in nume:
            if i != "(" and i != "'" and i != ")":
                if i == ',':
                    j += 1
                    break
                if j == 0:
                    name += i
        for i in boom:
            if i != "(" and i != "'" and i != ")":
                if i == ',':
                    j += 1
                    continue
                if j == 0:
                    line += i
                if j == 1:
                    collomn += i
                if j == 2:
                    break
            if i == ")":
                break

        line1 = ""
        collomn1 = ""
        k = 0
        m = 0
        for i in line:
            if i.isspace():
                print(i)
            else:
                line1 += i
        for i in collomn:
            if i.isspace():
                print(i)
            else:
                collomn1 += i
        h = ord(line1)
        c = ord(collomn1)
        while h > 48:
            k += 1
            h -= 1
        while c > 48:
            m += 1
            c -= 1
        print(name)
        book.EditareAtribut(name, k, m)
        print(k)
        # print(ord(line))
        print(m)
        return render_template("purchase.html")
    return render_template("edit.html", sheet=sheet)


@app.route("/add", methods=['GET', 'POST'])
def add():
    if request.method == 'POST' and 'ID' in request.form and 'name' in request.form and 'price' in request.form:
        id = request.form['ID']
        name = request.form['name']
        price = request.form['price']
        print(id)
        print(name)
        print(price)
        if book.Cautare_Dupa_Id(id) != -1:
            msg = 'Change Id, already exist !'
        inreg1 = [id, name, price]
        book.Adaugare_Inregistrare_Noua(inreg1)
    return render_template("add.html")


@app.route("/logout")
def logout():
    session.pop("user", None)
    return render_template("base.html")


@app.route("/books", methods=['GET', 'POST'])
def books():
    form = BasicForm()
    msg = ''
    bookEdit = load_workbook("booksDatabase.xlsx")
    sheet = bookEdit.active
    if form.validate_on_submit() and request.method == 'POST':
        name = form.ids.data
        print("name ", name)
        searchedAtribut = book.CautareAtributRow(name)
        print("searchedAtribut ", searchedAtribut)
        if searchedAtribut != -1:
            return render_template("bookSearched.html", sheet=sheet, variable=searchedAtribut)
    if request.method == 'POST' and 'buy' in request.form:
        buy = request.form['buy']

        if book.CautareAtribut(buy) != -1:
            valoare = book.CautareStoc(buy)
            print("Stoc carte ", valoare)
            if valoare != 0:
                Email.email_receiver = 'chintescu.bogdan@yahoo.com'
                WishList.insert({'Title': buy, 'Price': price})
            else:
                msg = 'Out of Stock !'
        else:
            msg = 'Not found'

        if buy == 'Buy':
            mail.Trimitere_Email(email_sender, email_password, 'Va multumim pentru achizite ', Email.email_receiver)
            return render_template("payment.html", variable=booksSize.books_value)

    return render_template("books.html", msg=msg, sheet=sheet, form=form)


@app.route("/search")
def bookSearched():
    book_edit = load_workbook("booksDatabase.xlsx")
    sheet = book_edit.active
    return render_template("bookSearched.html", sheet=sheet, variable=searchedAtribut)


@app.route("/purchase")
def purchase():
    return render_template("purchase.html", variable='value')


@app.route("/articles")
def articles():
    book_edit = load_workbook("articles.xlsx")
    sheet = book_edit.active
    return render_template("articles.html", sheet=sheet)


@app.route("/wishList", methods=['GET', 'POST'])
def wishList():
    elements = WishList.get_all()
    max_row = len(elements)
    if request.method == 'POST' and 'buy' in request.form:

        buy = request.form['buy']
        if buy == 'Price: 90':
            price = buy
            buy = 'All the Light We Cannot See'
            value = 90

        if buy == 'Price: 85':
            price = buy
            buy = 'Girl in Pieces'
            value = 85

        if book.CautareAtribut(buy) != -1:
            valoare = book.CautareStoc(buy)
            print("Stoc carte ", valoare)
            if valoare != 0:
                Email.email_receiver = 'chintescu.bogdan@yahoo.com'
                Book.insert({'Title': buy, 'Price': price})
                booksSize.books_value = booksSize.books_value + value
                book.Modificare_Valoare_In_Baza_Date(valoare, valoare - 1)
            else:
                msg = 'Out of Stock !'
        else:
            msg = 'Not found'

        if buy == 'Buy':
            mail.Trimitere_Email(email_sender, email_password, 'Va multumim pentru achizite ', Email.email_receiver)
            return render_template("payment.html", variable=booksSize.books_value)
    return render_template("wishList.html", max_row=max_row, elements=elements)

@app.route("/shoppingCart/", methods=['GET', 'POST'])
def shoppingCartRedirect():
    return redirect(url_for('shoppingCart'))

@app.route("/shoppingCart", methods=['GET', 'POST'])
def shoppingCart():
    global elements
    if request.method == 'POST' and 'buy' in request.form:
        buy = request.form['buy']
        if buy == 'Buy':
            Email.email_receiver = 'chintescu.bogdan@yahoo.com'
            mail.Trimitere_Email(email_sender, email_password, 'Va multumim pentru achizite ', Email.email_receiver)
            return render_template("payment.html", variable=booksSize.books_value)
    if request.method == 'POST' and 'emptyCart' in request.form:
        emptyCart = request.form['emptyCart']
        if emptyCart == 'EmptyCart':
            elements = []
            totalPrice = 0
            return render_template("shoppingCart.html", elements=elements, totalPrice=totalPrice)
    totalPrice = total_price(elements)
    return render_template("shoppingCart.html", elements=elements, totalPrice=totalPrice)


@app.route("/payment")
def paypal_payment():
    return render_template("payment.html", variable=value)


@app.route("/payments/<order_id>/capture", methods=["POST"])
def capture_payment(order_id):
    captured_payment = approve_payment(order_id)
    return jsonify(captured_payment)


def approve_payment(order_id):
    api_link = f"https://api-m.sandbox.paypal.com/v2/checkout/orders/%7Border_id%7D/capture"
    client_id = "ASByq-ekX4JSxdbttreMNlZlXShyIShRh3-WVEEOx7O6kq3YIG1_voD9VbaTASeVRolnCtCNHPwBgRtU"
    secret = ""
    basic_auth = HTTPBasicAuth(client_id, secret)
    headers = {
        "Content-Type": "application/json",
    }
    response = requests.post(url=api_link, headers=headers, auth=basic_auth)
    response.raise_for_status()
    json_data = response.json()
    return json_data

@app.route('/buy_book', methods=['POST'])
def buy_book():
    global elements
    data = request.get_json()
    title = data['title']
    price = data['price']
    elements.append({'title': title, 'price': price})
    return jsonify({'message': 'Book bought successfully!'})

def total_price(products):
    totalPrice = 0
    if products:
        for product in products:
            totalPrice = totalPrice + int(product['price'])
    return totalPrice


if __name__ == "__main__":
    Book = GenericModel('Book', ['Title', 'Price'])
    WishList = GenericModel('WishList', ['Title', 'Price'])
    app.run()
